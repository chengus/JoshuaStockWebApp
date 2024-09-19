from fastapi import FastAPI, Query, BackgroundTasks, Request
from fastapi.staticfiles import StaticFiles
import yfinance as yf
from datetime import timedelta
from datetime import datetime
from typing import List, Optional
import requests
import re
from requests.exceptions import RequestException
from fp.fp import FreeProxy
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import yagmail
from fastapi.responses import FileResponse, JSONResponse
import os, shutil
import time
import pandas as pd
import uuid
import asyncio
import aioschedule
from concurrent.futures import ProcessPoolExecutor
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

Errors = {}
LONG={}
SHORT={}
parent_etf={}
tradable_etfs=[]
t20holdings=[]
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:83.0) Gecko/20100101 Firefox/83.0",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
}
defualt_tickers = ['XLB', 'XLC', 'XLF', 'XLI', 'XLK', 'XLP', 'XLRE', 'XLU', 'XLV', 'XLY', 'XRT', 'SMH', 'QQQ', 'SPY', 'DIA', 'XME', 'XOP', 'ARKK', 'ARKG', 'BJK', 'CARZ', 'CLOU', 'GLD', 'IAI', 'IBB', 'IGV', 'IHE', 'ITA', 'ITB', 'IYT', 'IYW', 'KBE', 'KIE', 'KRE', 'KWEB', 'OIH', 'PEJ', 'QCLN', 'TLT', 'USO', 'VHT', 'VNQ', 'XBI']
processedTickers = []

app = FastAPI()
app.mount("/static", StaticFiles(directory="static"), name="static")

DOWNLOAD_DIR = ""
process_pool = ProcessPoolExecutor()

@app.get("/")
async def read_index():
    return FileResponse('static/index.html')

@app.get('/favicon.ico', include_in_schema=False)
async def favicon():
    return FileResponse("favicon.ico")

task_status = {}

async def startup_event():
    global DOWNLOAD_DIR
    asyncio.create_task(run_cleanup_scheduler())
    DOWNLOAD_DIR = create_download_directory()

async def run_cleanup_scheduler():
    aioschedule.every().day.at("00:00").do(cleanup_old_folders)
    while True:
        await aioschedule.run_pending()
        await asyncio.sleep(3600) 

async def cleanup_old_folders():
    current_time = datetime.now()
    for folder_name in os.listdir(DOWNLOAD_DIR):
        folder_path = os.path.join(DOWNLOAD_DIR, folder_name)
        if os.path.isdir(folder_path):
            folder_time = datetime.fromtimestamp(os.path.getctime(folder_path))
            if current_time - folder_time > timedelta(days=1):
                try:
                    shutil.rmtree(folder_path)
                    print(f"Deleted old folder: {folder_path}")
                except Exception as e:
                    print(f"Error deleting folder {folder_path}: {str(e)}")

@app.post('/run_screening')
async def register(
    background_tasks: BackgroundTasks, 
    sendEmailBool: Optional[bool] = Query(False, description="Whether to send an email"),
    tickersList: str = Query(..., description="Comma-separated list of tickers")
):
    logger.info(f"Received parameters: sendEmailBool={sendEmailBool}, tickersList={tickersList}")
    tickers = [ticker.strip() for ticker in tickersList.split(',')]
    task_id = str(uuid.uuid4())
    task_dir = os.path.join(DOWNLOAD_DIR, task_id)
    os.makedirs(task_dir, exist_ok=True)
    
    # Use ProcessPoolExecutor to run the screening task
    background_tasks.add_task(run_screening_task, tickers, sendEmailBool, task_id, task_dir)
    
    return JSONResponse(content={
        "message": "Screening started",
        "task_id": task_id,
        "tickers": tickers,
        "sendEmail": sendEmailBool
    })

@app.get('/task_status/{task_id}')
async def get_task_status(task_id: str):
    task_dir = os.path.join(DOWNLOAD_DIR, task_id)
    if not os.path.exists(task_dir):
        return JSONResponse(content={'status': 'not_found'})
    
    status_file = os.path.join(task_dir, 'status.txt')
    if os.path.exists(status_file):
        with open(status_file, 'r') as f:
            status = f.read().strip()
        return JSONResponse(content={'status': status})
    
    return JSONResponse(content={'status': 'processing'})

@app.get('/download/{task_id}')
async def download_excel(task_id: str, background_tasks: BackgroundTasks):
    task_dir = os.path.join(DOWNLOAD_DIR, task_id)
    if not os.path.exists(task_dir):
        return JSONResponse(content={"error": "File not found"}, status_code=404)
    
    excel_files = [f for f in os.listdir(task_dir) if f.endswith('.xlsx')]
    if not excel_files:
        return JSONResponse(content={"error": "Excel file not found"}, status_code=404)
    
    file_path = os.path.join(task_dir, excel_files[0])
    
    async def cleanup():
        await asyncio.sleep(5)  # Wait a bit to ensure file is sent
        try:
            shutil.rmtree(task_dir)
            print(f"Deleted folder after download: {task_dir}")
        except Exception as e:
            print(f"Error deleting folder {task_dir}: {str(e)}")

    background_tasks.add_task(cleanup)
    
    return FileResponse(
        file_path,
        filename=os.path.basename(file_path)
    )

@app.middleware("http")
async def log_requests(request: Request, call_next):
    # Log the request details
    logger.info(f"Request: {request.method} {request.url}")
    logger.info(f"Headers: {request.headers}")
    body = await request.body()
    logger.info(f"Body: {body.decode()}")
    
    response = await call_next(request)
    return response

def run_screening_task(tickers, sendEmailBool, task_id, task_dir):
    try:
        # Update status to processing
        with open(os.path.join(task_dir, 'status.txt'), 'w') as f:
            f.write('processing')
        
        # Run the actual screening process
        run_screening(tickers)
        file_path = export_to_excel(task_dir)
        
        if sendEmailBool:
            smtp_password = os.environ.get('SMTP_PSWD')
            if not smtp_password:
                raise ValueError("SMTP password not found in environment variables")
            send_email("1215.lucent@gmail.com", smtp_password, ["joshua@uasha.cn"])
        
        # Update status to completed
        with open(os.path.join(task_dir, 'status.txt'), 'w') as f:
            f.write('completed')
    except Exception as e:
        # Update status to error
        with open(os.path.join(task_dir, 'status.txt'), 'w') as f:
            f.write(f'error: {str(e)}')

def create_download_directory():
    download_dir = os.path.join(os.getcwd(), "download")
    if not os.path.exists(download_dir):
        os.makedirs(download_dir)
    return download_dir

def run_screening(tickers, donchian_range = 20):
    process_tickers(tickers, donchian_range)
    find_holdings()
    process_tickers(t20holdings, donchian_range)
    

def process_tickers(tickers, donchian_range):
    processedTickers.extend(tickers)
    if isinstance(tickers, str):
        tickers = [ticker.strip() for ticker in tickers.split(',')]
    for ticker in tickers:
        print(f"Processing {ticker}")
        Condition(ticker, donchian_range)

def Condition(ticker, window):
    try:
        df = yf.download(ticker, period='1y', interval='1wk')
        if df.empty:
            raise ValueError(f"No data found for ticker: {ticker}")
        data = df.iloc[::-1]

        # Calculate the Donchian Channel
        data['upper'] = data['High'].rolling(window).max().shift(-window)
        data['lower'] = data['Low'].rolling(window).min().shift(-window)
        data['mid'] = (data['upper'] + data['lower']) / 2

        # Drop the rows with NaN values which are present at the start of the dataframe
        data = data.dropna()

        # Adjust the logic to check the condition within a specific date range
        results = {"LONG": None, "SHORT": None}
        earliest_date = None
        for i in range(0, 3):
            current_week = data.iloc[i]
            previous_5_weeks = data.iloc[i + 1:i + 6]

            # Check if the middle band is within the range of the highs and lows of the previous 5 weeks
            mid_within_range = (previous_5_weeks['Low'] <= previous_5_weeks['mid']).any() and (previous_5_weeks['High'] >= previous_5_weeks['mid']).any()

            weekLong = mid_within_range and (current_week['Close'] > current_week['mid']) and (current_week['Close'] > current_week['Open'])
            weekShort = mid_within_range and (current_week['Close'] < current_week['mid']) and (current_week['Close'] < current_week['Open'])

            if weekLong or weekShort:
                start_date = data.index[i]
                position_type = "LONG" if weekLong else "SHORT"
                results[position_type] = start_date
                if earliest_date is None or start_date < earliest_date:
                    earliest_date = start_date

        if earliest_date is not None:
            earliest_date = earliest_date - timedelta(days=90)
            df = yf.download(ticker, start=earliest_date.strftime('%Y-%m-%d'), interval='1d')
            daily_data = df.iloc[::-1]

            # Calculate the Donchian Channel for daily data
            daily_data['upper'] = daily_data['High'].rolling(window).max().shift(-window)
            daily_data['lower'] = daily_data['Low'].rolling(window).min().shift(-window)
            daily_data['mid'] = (daily_data['upper'] + daily_data['lower']) / 2

            for position_type, signal_date in results.items():
                if signal_date:
                    signal_data = daily_data.loc[:signal_date]
                    signal_dates = signal_data[(signal_data['Close'] > signal_data['upper'] if position_type == "LONG" else signal_data['Close'] < signal_data['lower']) & 
                                               (signal_data['Close'] > signal_data['Open'] if position_type == "LONG" else signal_data['Close'] < signal_data['Open'])].index

                    if not signal_dates.empty:
                        target_dict = LONG if position_type == "LONG" else SHORT
                        for date in signal_dates:
                            most_recent_close = round(daily_data.iloc[0]['Close'], 2)
                            signal_close = round(signal_data.loc[date]['Close'], 2)
                            weekly_donchian_mid = round(signal_data.loc[date]['mid'], 2)
                            pct_difference = round(((most_recent_close / signal_close) * 100) - 100, 2)
                            wdm_pct_difference = round(((signal_close / weekly_donchian_mid) * 100) - 100, 2)
                            
                            new_entry = [
                                date.strftime('%Y-%m-%d'),
                                str(signal_close),
                                str(most_recent_close),
                                str(pct_difference) + "%",
                                position_type,
                                str(weekly_donchian_mid),
                                str(wdm_pct_difference) + "%"
                            ]
                            
                            if ticker in target_dict:
                                target_dict[ticker].append(new_entry)
                            else:
                                target_dict[ticker] = [new_entry]
        
        # Mark repeated entries
        for key, entries in LONG.items():
            if len(entries) > 1:
                for i in range(len(entries)-1):
                    entries[i].append("R")
                entries[-1].append("")
            else:
                entries[0].append("")

        for key, entries in SHORT.items():
            if len(entries) > 1:
                for i in range(len(entries)-1):
                    entries[i].append("R")
                entries[-1].append("")
            else:
                entries[0].append("")

        print("LONG:")
        print(LONG)
        print()
        print("SHORT:")
        print(SHORT)

    except Exception as e:
        print(f"Error processing {ticker}: {e}")
        Errors[ticker] = str(e)

def find_holdings():
    tradable_etfs = list(LONG.keys()) + list(SHORT.keys())
    print("Tradable ETFs:")
    print(tradable_etfs)
    for etf in tradable_etfs:
        holdings = get_etf_holdings(etf)
        parent_etf[etf] = holdings
        t20holdings.extend(holdings)
        print(t20holdings)
        processedTickers.extend(holdings)

def get_etf_holdings(etf, max_retries=99):
    url = f"https://www.zacks.com/funds/etf/{etf}/holding"

    for attempt in range(max_retries):
        try:
            proxy = FreeProxy(country_id=['US', 'GB']).get()
            print(f"Using proxy: {proxy}")

            with requests.Session() as req:
                req.headers.update(headers)
                req.proxies = {'http': proxy, 'https': proxy}
                r = req.get(url, timeout=10, verify=False)
                print(f"Extracting: {r.url}")
                goal = re.findall(r'etf\\\/(.*?)\\', r.text)
                goal = goal[:19]
                return goal

        except RequestException as e:
            print(f"Attempt {attempt + 1} failed: {str(e)}")
            if attempt == max_retries - 1:
                print("Max retries reached. Unable to fetch data.")
                return None

    return None


def export_to_excel(task_dir):
    global file_path, file_name
    try:
        file_name = f"{datetime.today().strftime('%Y-%m-%d')}_Stock_Screening_Results.xlsx"
        file_path = os.path.join(task_dir, file_name)
        
        if os.path.exists(file_path):
            os.remove(file_path)

        def find_parent_etf(ticker):
            if ticker in parent_etf:
                return ticker
            for etf, holdings in parent_etf.items():
                if ticker in holdings:
                    return etf
            return None
        
        def format_excel(writer, sheet_name):
            workbook = writer.book
            worksheet = workbook[sheet_name]
            
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.font = Font(name='DengXian', size=12)
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer: # pylint: disable=abstract-class-instantiated 
            # LONG data
            long_data = []
            for ticker, records in LONG.items():
                for record in records:
                    parent = find_parent_etf(ticker)
                    long_data.append([ticker, parent] + list(record))
            long_df = pd.DataFrame(long_data, columns=["Ticker", "Parent ETF", "Breakout Date", "Breakout Closing Price", "Last Daily Price", "% Difference", "L/S", "Weekly Donchian Mid", "WDM % Difference","Repeated?"])
            long_df.to_excel(writer, sheet_name='LONG', index=False)
            format_excel(writer, 'LONG')

            # SHORT data
            short_data = []
            for ticker, records in SHORT.items():
                for record in records:
                    parent = find_parent_etf(ticker)
                    short_data.append([ticker, parent] + list(record))
            short_df = pd.DataFrame(short_data, columns=["Ticker", "Parent ETF", "Breakout Date", "Breakout Closing Price", "Last Daily Price", "% Difference", "L/S", "Weekly Donchian Mid", "WDM % Difference","Repeated?"])
            short_df.to_excel(writer, sheet_name='SHORT', index=False)
            format_excel(writer, 'SHORT')

        print(f"Data exported to Excel successfully: {file_path}")
        return file_path
    except Exception as e:
        print(f"An error occurred while exporting to Excel: {str(e)}")
        raise


def send_email(sender_email, app_password, recipient_email):
    # Initialize yagmail SMTP client
    yag = yagmail.SMTP(sender_email, app_password)

    def find_parent_etf(ticker):
        # If the ticker is an ETF itself, it is its own parent
        if ticker in parent_etf:
            return ticker
        # Otherwise, find which ETF the ticker belongs to
        for etf, holdings in parent_etf.items():
            if ticker in holdings:
                return etf
        return None

    # Email content
    subject = "Daily Stock Screening Service"
    
    # Generate HTML content for Long and Short tables
    html_content = "<h2>Trade Signals</h2>"
    if LONG:
        html_content += "<h3>Long Positions:</h3>"
        html_content += "<table border='1'><tr><th>Ticker</th><th>Parent ETF</th><th>Breakout Date</th><th>Breakout Closing Price</th><th>Most Recent Close</th><th>Percentage Difference</th></tr>"
        for ticker, entries in LONG.items():
            parent = find_parent_etf(ticker)
            for entry in entries:
                html_content += f"<tr><td>{ticker}</td><td>{parent}</td><td>{entry[0]}</td><td>{entry[1]}</td><td>{entry[2]}</td><td>{entry[3]}</td></tr>"
        html_content += "</table>"

    if SHORT:
        html_content += "<h3>Short Positions:</h3>"
        html_content += "<table border='1'><tr><th>Ticker</th><th>Parent ETF</th><th>Breakout Date</th><th>Breakout Closing Price</th><th>Most Recent Close</th><th>Percentage Difference</th></tr>"
        for ticker, entries in SHORT.items():
            parent = find_parent_etf(ticker)
            for entry in entries:
                html_content += f"<tr><td>{ticker}</td><td>{parent}</td><td>{entry[0]}</td><td>{entry[1]}</td><td>{entry[2]}</td><td>{entry[3]}</td></tr>"
        html_content += "</table>"

    # Adding the list of all processed tickers
    html_content += "<h3>Processed Tickers:</h3>"
    html_content += "<ul>"
    for ticker in processedTickers:
        html_content += f"<li>{ticker}</li>"
    html_content += "</ul>"

    html_content += "<h3>Errors:</h3>"
    html_content += "<ul>"
    for ticker, error in Errors.items():
        html_content += f"<li>{ticker}: {error}</li>"
    html_content += "</ul>"

    while not os.path.exists(file_path):
        print("File does not exist")
        time.sleep(1)

    # Send the email
    try:
        yag.send(
            to=recipient_email,
            subject=subject,
            contents=html_content,
            attachments=file_path
        )
        print("Email sent successfully")
    except Exception as e:
        print(f"Error sending email: {e}")
